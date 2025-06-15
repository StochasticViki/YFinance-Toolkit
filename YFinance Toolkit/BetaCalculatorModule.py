import yfinance as yf
import numpy as np
import requests
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import to_excel
import sys
import os
from datetime import datetime, timedelta

max_tries = 4

status = True
# Getting end date then taking the year and subtracting 5 from it and then making it start date
if getattr(sys, 'frozen', False):
    # Running as a bundled executable
    # The _MEI_ path is where PyInstaller unpacks the files
    base_path = sys._MEIPASS # Use _MEI_TEMP if you use --temp-dir
    # If not using --temp-dir, use sys._MEI_
            # base_path = sys._MEI_
else:
    # Running as a regular Python script
    base_path = os.path.dirname(os.path.abspath(__file__))
        


def generate_beta(ticker, start_date, end_date):
    ticker.insert(0,"^BSESN")
    intrvl = "1d"
    columns = ["Close"]
    quotes = []
    blankcolumns = []

    # Convert to datetime
    todate_obj = datetime.strptime(end_date, "%Y-%m-%d")
    fromdate_obj = datetime.strptime(start_date, "%Y-%m-%d")
    # Get next 5 dates
    todate_after = todate_obj + timedelta(days=1)
    end_date = todate_after.strftime("%Y-%m-%d")
    

    # this creates empty columns
    def create_empty_dataframe(index):
        return pd.DataFrame(columns=[" "], index=index)

    # Downloading each ticker's CLose data an storing it in quotes, also noting no. of tickers in multiples of 4
    for i, t in enumerate(ticker):
        attempts = 0
        success = False
        global max_tries
        while attempts < max_tries and not success:
            try:
                data = yf.download(ticker[i], start=start_date, end=end_date, interval=intrvl)[columns]
                if data.empty:
                    print(f"Warning: No data found for {ticker[i]}")
                    attempts += 1
                else:
                    quotes.append(data)
                    success = True
            except Exception as e:
                print(f"Error downloading {ticker[i]}: {e}")
                attempts += 1
                
            if not success and attempts >= max_tries:
                quotes.append(pd.DataFrame(columns=["Close"]))

            blankcolumns.append((i*4)+4) 
    max_tries = 4
    combined = pd.concat(quotes,axis=1)
    combined = pd.DataFrame(data=combined)
    # combined = combined.reset_index()
    combined = combined.sort_index(ascending=False)
    # combined["Date"] = pd.to_datetime(combined["Date"])
    # combined.sort_values(by="Date", ascending=False)
    cut_off_date = to_excel(todate_obj)
    
    filename = "_".join([t.replace('.', '') for t in ticker])
    combined.to_excel(f"{base_path}/{filename}.xlsx", sheet_name="data")

    wb = openpyxl.load_workbook(f"{base_path}/{filename}.xlsx")
    sheet = wb.active
    merged_ranges = list(sheet.merged_cells.ranges)

    for merged_range in merged_ranges:
        sheet.unmerge_cells(str(merged_range))
    

    sheet.delete_rows(2, 2)

    sheet.insert_rows(1,5)
    sheet.cell(row=2, column=1).value = "Beta Calculation"

    center_alignment = Alignment(horizontal='center', vertical='center')
    dateformat = 'DD-MMM-YY'
    thin_border_side = Side(border_style="thin", color="000000")
    border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side
    )
    bold_font = Font(bold=True)
    Underline = Border(bottom=Side(border_style='medium'))
    sheet['A2'].font = bold_font

    sheet.insert_cols(1,1)
    sheet.column_dimensions['A'].width = 3

    def get_last_column(sheet):
        max_column = 1
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    max_column = max(max_column, cell.column)
        return max_column

    fill_blue = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    fill_lightblue = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    fill_error = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")

    for colum in range(2, 1+get_last_column(sheet)):
        sheet.cell(row=6, column=colum).fill = fill_blue

    for n in blankcolumns:
        sheet.insert_cols(n,3)

    for column in blankcolumns:
        blank_counter = 0
        for row in reversed(range(7,sheet.max_row)):
            if sheet.cell(row=row,column= column-1 ).value == None:
                sheet.cell(row=row,column= column-1 ).fill = fill_error
                blank_counter += 1
                sheet.cell(row=row,column= column-1 ).value = sheet.cell(row=row+1,column= column-1 ).value
            else: pass
            sheet.cell(row=5, column=column-1).value = blank_counter
            sheet.cell(row=5, column=column-2).value = "Empty Dates"            
                 

    def excel_reference(row, col):
        colno= sheet.cell(row=row, column=col).column
        col_letter = get_column_letter(colno)
        # Combine the column letter and row number to form the reference
        return f"{col_letter}{row}"


    start_row = 7  # Assuming headers are in row 6, data starts from row 7
    end_row = sheet.max_row

    for column in blankcolumns:
        for row in range(start_row, end_row):
            currentcell = excel_reference(row, column-1)
            prevcell = excel_reference(row+1, column-1)

            if currentcell is not None and prevcell is not None:
                try:
                    sheet.cell(row=row, column=column).value = f"={currentcell}/{prevcell}-1"
                    sheet.cell(row=row, column=column).number_format = '0.00%'
                except ZeroDivisionError:
                    sheet.cell(row=row, column=column).value = None
            else:
                sheet.cell(row=row, column=column).value = None


    no_border = Border(left=Side(border_style=None),
                    right=Side(border_style=None),
                    top=Side(border_style=None),
                    bottom=Side(border_style=None))

    for row in range(7,sheet.max_row+1):
        sheet.cell(row=row, column= 2).border = no_border

    xvalues = []
    for change in range(8,sheet.max_row+1):
        xvalues.append(sheet.cell(row=change, column=4).value)

    bseref = f"D7:{excel_reference(sheet.max_row-1,4)}"

    for n in blankcolumns:
        
        if n == 4:
            sheet.cell(row=6, column= n-1).alignment = center_alignment
            sheet.cell(row=6, column= n).alignment = center_alignment
            pass
        else: 
            sheet.cell(row=6, column= n-2).value = "Date"
            sheet.cell(row=6, column= n-2).fill = fill_blue
            sheet.cell(row=6, column= n-2).border = border
            sheet.cell(row=6, column= n-2).font = bold_font
            sheet.cell(row=6, column= n-2).alignment = center_alignment

            sheet.cell(row=6, column= n-1).value = "Close"
            sheet.cell(row=6, column= n-1).fill = fill_blue
            sheet.cell(row=6, column= n-1).border = border
            sheet.cell(row=6, column= n-1).font = bold_font
            sheet.cell(row=6, column= n-1).alignment = center_alignment


            sheet.cell(row=4, column= n-2).value = "SLOPE"
            sheet.cell(row=4, column= n-2).font = bold_font
            sheet.cell(row=4, column= n).font = bold_font

            sheet.cell(row=6, column= n).value = "% change"
            sheet.cell(row=6, column= n).font = bold_font
            sheet.cell(row=6, column= n).border = border
            sheet.cell(row=6, column= n).fill = fill_blue
            sheet.cell(row=6, column= n).alignment = center_alignment

            
            
            firstcell = excel_reference(7,n)
            lastcell = excel_reference(sheet.max_row-1,n)
            ref= f"=SLOPE({firstcell}:{lastcell},{bseref})"

            sheet.cell(row=4, column= n).value = ref
            sheet.cell(row=4, column= n).fill = fill_lightblue
            sheet.cell(row=4, column= n).border = border
            sheet.cell(row=4, column= n).number_format = "0.00"


    sheet.cell(row=6, column= 4).value = "% change"
    sheet.cell(row=6, column= 4).font = bold_font
    sheet.cell(row=6, column= 4).border = border
    sheet.cell(row=6, column= 4).fill = fill_blue

    for n in blankcolumns:
        if n == 4:
            pass
        else:
            sheet.column_dimensions[sheet.cell(row=1, column= n-1).column_letter].width = 13
            for row in range(7,sheet.max_row+1):
                sheet.cell(row=row, column= n-2).value = sheet.cell(row=row, column= 2).value
                sheet.cell(row=row, column= n-2).number_format = 'DD-MMM-YY' 
                sheet.cell(row=row, column= n-1).number_format = '0.00'

    for row in range(7,sheet.max_row+1):
        sheet.cell(row=row, column= 2).number_format = 'DD-MMM-YY'
        sheet.cell(row=row, column= 2).font = Font(bold=False)

    sheet.sheet_view.showGridLines = False

    # blankcolumns.append(blankcolumns[-1]+2)

    for n, ind in zip(blankcolumns,ticker):
        if n == 4:
            pass 
        else: 
            sheet.cell(row= 3, column= n-2).value = ind.upper()
            sheet.cell(row= 3, column= n-2).font = bold_font  

    sheet.row_dimensions[3].height = 26
    sheet.cell(row=3,column=2).value = "S&P BSE Sensex"
    sheet.cell(row=3,column=2).font = bold_font

    for n in blankcolumns:
        sheet.column_dimensions[sheet.cell(row=1, column= n-3).column_letter].width = 6
        sheet.column_dimensions[sheet.cell(row=1, column= n-2).column_letter].width = 13
        sheet.column_dimensions[sheet.cell(row=1, column= n-1).column_letter].width = 13
        sheet.column_dimensions[sheet.cell(row=1, column= n).column_letter].width = 13
        

    for colum in range(2, 5):
        sheet.cell(row=2, column= colum).border = Underline
    
    del ticker[0]

    wb.save(f"{base_path}/BetaReport.xlsx")

def generate_volatility(ticker, start_date, end_date):
    intrvl = "1d"
    columns = ["Close"]
    quotes = []
    blankcolumns = []

    # Convert to datetime
    todate_obj = datetime.strptime(end_date, "%Y-%m-%d")
    fromdate_obj = datetime.strptime(start_date, "%Y-%m-%d")
    # Get next 5 dates
    todate_after = todate_obj + timedelta(days=1)
    end_date = todate_after.strftime("%Y-%m-%d")
    

    # this creates empty columns
    def create_empty_dataframe(index):
        return pd.DataFrame(columns=[" "], index=index)

    # Downloading each ticker's CLose data an storing it in quotes, also noting no. of tickers in multiples of 4
    for i, t in enumerate(ticker):
        attempts = 0
        success = False
        global max_tries
        while attempts < max_tries and not success:
            try:
                data = yf.download(ticker[i], start=start_date, end=end_date, interval=intrvl)[columns]
                if data.empty:
                    print(f"Warning: No data found for {ticker[i]}")
                    attempts += 1
                else:
                    quotes.append(data)
                    success = True
            except Exception as e:
                print(f"Error downloading {ticker[i]}: {e}")
                attempts += 1
                
            if not success and attempts >= max_tries:
                quotes.append(pd.DataFrame(columns=["Close"]))

            blankcolumns.append((i*4)+4) 
    max_tries = 4
    combined = pd.concat(quotes,axis=1)
    combined = pd.DataFrame(data=combined)
    # combined = combined.reset_index()
    combined = combined.sort_index(ascending=False)
    # combined["Date"] = pd.to_datetime(combined["Date"])
    # combined.sort_values(by="Date", ascending=False)
    cut_off_date = to_excel(todate_obj)
    
    filename = "_".join([t.replace('.', '') for t in ticker])
    combined.to_excel(f"{base_path}/{filename}.xlsx", sheet_name="data")

    wb = openpyxl.load_workbook(f"{base_path}/{filename}.xlsx")
    sheet = wb.active
    merged_ranges = list(sheet.merged_cells.ranges)

    for merged_range in merged_ranges:
        sheet.unmerge_cells(str(merged_range))
    

    sheet.delete_rows(2, 2)

    sheet.insert_rows(1,5)
    sheet.cell(row=2, column=1).value = "Volatility Calculation"

    center_alignment = Alignment(horizontal='center', vertical='center')
    dateformat = 'DD-MMM-YY'
    thin_border_side = Side(border_style="thin", color="000000")
    border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side
    )
    bold_font = Font(bold=True)
    Underline = Border(bottom=Side(border_style='medium'))
    sheet['A2'].font = bold_font

    sheet.insert_cols(1,1)
    sheet.column_dimensions['A'].width = 3

    def get_last_column(sheet):
        max_column = 1
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    max_column = max(max_column, cell.column)
        return max_column

    fill_blue = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    fill_lightblue = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    fill_error = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")

    for colum in range(2, 1+get_last_column(sheet)):
        sheet.cell(row=6, column=colum).fill = fill_blue

    for n in blankcolumns:
        sheet.insert_cols(n,3)

    for column in blankcolumns:
        blank_counter = 0
        for row in reversed(range(7,sheet.max_row)):
            if sheet.cell(row=row,column= column-1 ).value == None:
                sheet.cell(row=row,column= column-1 ).fill = fill_error
                blank_counter += 1
                sheet.cell(row=row,column= column-1 ).value = sheet.cell(row=row+1,column= column-1 ).value
            else: pass
            sheet.cell(row=5, column=column-1).value = blank_counter
            sheet.cell(row=5, column=column-2).value = "Empty Dates"            
                 

    def excel_reference(row, col):
        colno= sheet.cell(row=row, column=col).column
        col_letter = get_column_letter(colno)
        # Combine the column letter and row number to form the reference
        return f"{col_letter}{row}"


    start_row = 7  # Assuming headers are in row 6, data starts from row 7
    end_row = sheet.max_row

    for column in blankcolumns:
        for row in range(start_row, end_row):
            currentcell = excel_reference(row, column-1)
            prevcell = excel_reference(row+1, column-1)

            if currentcell is not None and prevcell is not None:
                try:
                    sheet.cell(row=row, column=column).value = f"={currentcell}/{prevcell}-1"
                    sheet.cell(row=row, column=column).number_format = '0.00%'
                except ZeroDivisionError:
                    sheet.cell(row=row, column=column).value = None
            else:
                sheet.cell(row=row, column=column).value = None


    no_border = Border(left=Side(border_style=None),
                    right=Side(border_style=None),
                    top=Side(border_style=None),
                    bottom=Side(border_style=None))

    for row in range(7,sheet.max_row+1):
        sheet.cell(row=row, column= 2).border = no_border

    xvalues = []
    for change in range(8,sheet.max_row+1):
        xvalues.append(sheet.cell(row=change, column=4).value)

    bseref = f"D7:{excel_reference(sheet.max_row-1,4)}"

    for n in blankcolumns:
        
         
            sheet.cell(row=6, column= n-2).value = "Date"
            sheet.cell(row=6, column= n-2).fill = fill_blue
            sheet.cell(row=6, column= n-2).border = border
            sheet.cell(row=6, column= n-2).font = bold_font
            sheet.cell(row=6, column= n-2).alignment = center_alignment

            sheet.cell(row=6, column= n-1).value = "Close"
            sheet.cell(row=6, column= n-1).fill = fill_blue
            sheet.cell(row=6, column= n-1).border = border
            sheet.cell(row=6, column= n-1).font = bold_font
            sheet.cell(row=6, column= n-1).alignment = center_alignment


            sheet.cell(row=4, column= n-2).value = "Sigma"
            sheet.cell(row=4, column= n-2).font = bold_font
            sheet.cell(row=4, column= n).font = bold_font

            sheet.cell(row=6, column= n).value = "% change"
            sheet.cell(row=6, column= n).font = bold_font
            sheet.cell(row=6, column= n).border = border
            sheet.cell(row=6, column= n).fill = fill_blue
            sheet.cell(row=6, column= n).alignment = center_alignment

            
            
            firstcell = excel_reference(7,n)
            lastcell = excel_reference(sheet.max_row-1,n)
            ref= f"=(STDEV({firstcell}:{lastcell})*SQRT(252))"

            sheet.cell(row=4, column= n).value = ref
            sheet.cell(row=4, column= n).data_type = "f"
            sheet.cell(row=4, column= n).fill = fill_lightblue
            sheet.cell(row=4, column= n).border = border
            sheet.cell(row=4, column= n).number_format = "0.00"


    sheet.cell(row=6, column= 4).value = "% change"
    sheet.cell(row=6, column= 4).font = bold_font
    sheet.cell(row=6, column= 4).border = border
    sheet.cell(row=6, column= 4).fill = fill_blue

    for n in blankcolumns:
            sheet.column_dimensions[sheet.cell(row=1, column= n-1).column_letter].width = 13
            for row in range(7,sheet.max_row+1):
                sheet.cell(row=row, column= n-2).value = sheet.cell(row=row, column= 2).value
                sheet.cell(row=row, column= n-2).number_format = 'DD-MMM-YY' 
                sheet.cell(row=row, column= n-1).number_format = '0.00'

    for row in range(7,sheet.max_row+1):
        sheet.cell(row=row, column= 2).number_format = 'DD-MMM-YY'
        sheet.cell(row=row, column= 2).font = Font(bold=False)

    sheet.sheet_view.showGridLines = False

    # blankcolumns.append(blankcolumns[-1]+2)

    for n, ind in zip(blankcolumns,ticker):
        sheet.cell(row= 3, column= n-2).value = ind.upper()
        sheet.cell(row= 3, column= n-2).font = bold_font  

    sheet.row_dimensions[3].height = 26
    # sheet.cell(row=3,column=2).value = "S&P BSE Sensex"
    sheet.cell(row=3,column=2).font = bold_font

    for n in blankcolumns:
        sheet.column_dimensions[sheet.cell(row=1, column= n-3).column_letter].width = 6
        sheet.column_dimensions[sheet.cell(row=1, column= n-2).column_letter].width = 13
        sheet.column_dimensions[sheet.cell(row=1, column= n-1).column_letter].width = 13
        sheet.column_dimensions[sheet.cell(row=1, column= n).column_letter].width = 13
        

    for colum in range(2, 5):
        sheet.cell(row=2, column= colum).border = Underline

    wb.save(f"{base_path}/VolatilityReport.xlsx")

def generate_prices(ticker, start_date, end_date, columns):
    intrvl = "1d"
    
    quotes = []
    blankcolumns = []

    # Convert to datetime
    todate_obj = datetime.strptime(end_date, "%Y-%m-%d")
    fromdate_obj = datetime.strptime(start_date, "%Y-%m-%d")
    # Get next 5 dates
    todate_after = todate_obj + timedelta(days=1)
    end_date = todate_after.strftime("%Y-%m-%d")
    

    # this creates empty columns
    def create_empty_dataframe(index):
        return pd.DataFrame(columns=[" "], index=index)

    # Downloading each ticker's CLose data an storing it in quotes, also noting no. of tickers in multiples of 4
    for i, t in enumerate(ticker):
        attempts = 0
        success = False
        global max_tries
        while attempts < max_tries and not success:
            try:
                data = yf.download(ticker[i], start=start_date, end=end_date, interval=intrvl)[columns]
                if data.empty:
                    print(f"Warning: No data found for {ticker[i]}")
                    attempts += 1
                else:
                    quotes.append(data)
                    success = True
            except Exception as e:
                print(f"Error downloading {ticker[i]}: {e}")
                attempts += 1
                
            if not success and attempts >= max_tries:
                quotes.append(pd.DataFrame(columns=["Close"]))

            blankcolumns.append((i*4)+4) 
    max_tries = 4
    combined = pd.concat(quotes,axis=1)
    combined = pd.DataFrame(data=combined)
    # combined = combined.reset_index()
    combined = combined.sort_index(ascending=False)
    # combined["Date"] = pd.to_datetime(combined["Date"])
    # combined.sort_values(by="Date", ascending=False)
    cut_off_date = to_excel(todate_obj)
    
    filename = "_".join([t.replace('.', '') for t in ticker])
    combined.to_excel(f"{base_path}/{filename}.xlsx", sheet_name="data")


# generate_volatility(["reliance.ns","aapl","msft"], "2022-03-31", "2025-03-31")